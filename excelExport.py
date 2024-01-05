import shutil

import numpy as np
import os
import openpyxl
import win32com.client
from PIL import Image

list_of_colors = [[118, 22, 23],
                  [160, 57, 25],
                  [210, 105, 30],
                  [227, 155, 78],
                  [244,199,120],
                  [255, 229, 149],
                  [204, 255, 102],
                  [118, 238, 0],
                  [0, 191, 0],
                  [0, 127, 0],
                  [0, 77, 0],
                  [4, 25, 67],
                  [0, 0, 140],
                  [0, 0, 242],




]


colorValues=[
    .1,
    .3,
    .5,
    .7,
    .85,
    .95,
    1.05,
    1.175,
    1.375,
    1.75,
    2.5,
    3.5,
    5,
    7
]
color = [155,155,155]

mycolumns={
    "pdays90" : 0,
    "pdays30" : 1,
    "pdays15": 2,
    "fdays15": 3,
    "fmonth1": 4,
    "fmonth2": 5,
    "fmonth3": 6,
    "fmonth4": 7
}

class keys:
    def __init__(self):
        self.colors_keys=[0 for col in list_of_colors]
        self.total=0


def closest(colors,color):
    colors = np.array(colors)
    color = np.array(color)
    distances = np.sqrt(np.sum((colors-color)**2,axis=1))
    index_of_smallest = np.where(distances==np.amin(distances))
    smallest_distance = colors[index_of_smallest]
    return smallest_distance


def  indexof(array,val):
    for i, city in enumerate(array):
        if ((city == val).all()):
            return i
    return -1


# Press the green button in the gutter to run the script.
def ExportImagestoExcel(powerpointFileAddress,pathExcel,pathImages):




    if not os.path.exists(pathImages):
        os.mkdir(pathImages)
    else:
        for file in os.listdir(pathImages):
            os.remove(os.path.join(pathImages, file))


    wk = openpyxl.load_workbook(pathExcel)
    sh = wk["Settings"]
    rowNumber=sh.cell(1,2).value
    columnNumber = sh.cell(2, 2).value
    sh = wk["Analysis"]




    ppt = win32com.client.Dispatch('PowerPoint.Application')
    ppt.visible = True

    ppres = ppt.Presentations.Open(powerpointFileAddress, False)
    ppt.Run(f"{os.path.basename(powerpointFileAddress)}!ExportAllImages.exportAllSubImages")

    try:
        ppres.close()
        ppt.quit()
    except:
        pass

    for file in os.listdir(pathImages+"\\"):
        # print(os.path.join(pathImages,file))
        photo = Image.open(os.path.join(pathImages,file))  # your image
        photo = photo.convert('RGB')

        width = photo.size[0]  # define W and H
        height = photo.size[1]
        col= keys()

        for y in range(0, height):  # each pixel has coordinates
            row = ""
            for x in range(0, width):
                RGB = photo.getpixel((x, y))
                r,g,b=RGB
                if not (r>254  and g>254 and b >254 ) and not (r<2  and g<2 and b <2 ):
                    # print(closest(colors=list_of_colors,color=RGB),indexof(list_of_colors,closest(colors=list_of_colors,color=RGB)))
                    col.colors_keys[indexof(list_of_colors,closest(colors=list_of_colors,color=RGB))]+=1
                    col.total+=1

                    # print(closest(colors=list_of_colors,color=RGB))



        ratios=[x/col.total for x in col.colors_keys] if col.total>0 else col.colors_keys

        finalvalue=0
        for index,rat in enumerate(ratios):
            finalvalue+=rat*colorValues[index]
        arr = file.split("@")
        arr[1]=arr[1].replace(".png","")
        arr[1]=arr[1].lower().replace("_","").replace(" ","")
        row=int(arr[0])-1
        col=mycolumns[arr[1]]

        sh.cell(rowNumber+row,columnNumber+col).value=round(finalvalue*100,2)



    wk.save(pathExcel)
    wk.close()

    # if os.path.exists(pathExcel):
    #     os.remove(pathExcel)


    # closest_color = closest(list_of_colors, color)
    # print(closest_color)

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
    os.system("TASKKILL /F /IM powerpnt.exe")
